"""
Parse all WB 2021 Statistical Report XLSX files into clean CSVs.

Inputs: data/electoral_history/eci_detailed/set_A/*.xlsx
Outputs: data/electoral_history/2021/*.csv
"""
from openpyxl import load_workbook
import csv, os, re

SRC = '/Users/atharvapandey/Kairosity/WBData/data/electoral_history/eci_detailed/set_A'
DST = '/Users/atharvapandey/Kairosity/WBData/data/electoral_history/2021'
os.makedirs(DST, exist_ok=True)


def xlsx_to_csv(xlsx_path, csv_path, skip_rows=0):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Find header row — first row with >1 non-empty cell
    start = skip_rows
    with open(csv_path, 'w', newline='') as f:
        w = csv.writer(f)
        for r in rows[start:]:
            vals = ['' if v is None else str(v) for v in r]
            if any(v.strip() for v in vals):
                w.writerow(vals)
    return len(rows[start:])


FILES = [
    ('1-Other Abbreviations And Description.xlsx', 'abbreviations.csv', 1),
    ('2-List of Successful Candidates.xlsx', 'successful_candidates.csv', 1),
    ('3-List Of Political Parties Participated.xlsx', 'parties_participated.csv', 1),
    ('4-Highlight.xlsx', 'highlights.csv', 1),
    ('5-Performance of Political Parties.xlsx', 'party_performance.csv', 1),
    ('6-Electors Data Summary.xlsx', 'electors_summary.csv', 1),
    ('7-Individual Performance Of Women Candidate.xlsx', 'women_candidates.csv', 1),
    ('8-Constituency Data Summary.xlsx', 'constituency_summary.csv', 1),
    ('9-Candidate Data Summary.xlsx', 'candidate_summary.csv', 1),
    ('10-Detailed Results.xlsx', 'detailed_results.csv', 1),
    ('11-Electors Data Summary Annxure.xlsx', 'electors_annexure.csv', 1),
]

for src_name, dst_name, skip in FILES:
    src = f'{SRC}/{src_name}'
    dst = f'{DST}/{dst_name}'
    if not os.path.exists(src):
        print(f"  ✗ missing: {src_name}")
        continue
    n = xlsx_to_csv(src, dst, skip)
    sz = os.path.getsize(dst)
    print(f"  ✓ {dst_name}: {n} rows, {sz//1024} KB")

print(f"\nAll CSVs in: {DST}")
